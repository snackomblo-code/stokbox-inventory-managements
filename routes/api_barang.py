"""API CRUD barang (item inventaris)."""
from __future__ import annotations

import io as _io

from fenrir import (
    Body,
    Blueprint,
    File,
    HTTPBadRequest,
    HTTPNotFound,
    Query,
    Response,
    UploadFile,
    request,
)
from openpyxl import Workbook, load_workbook

from services import barang_service, cloudinary_service
from utils.decorators import api_login_required, role_required

barang_bp = Blueprint("api-barang", url_prefix="/api/barang")


@barang_bp.get("")
@api_login_required
async def index(
    keyword: str = Query(""),
    kategori_id: str = Query(""),
    stok: str = Query(""),
):
    return {
        "data": barang_service.list_barang(
            keyword=keyword, kategori_id=kategori_id, stok_filter=stok
        )
    }


@barang_bp.get("/check-kode")
@api_login_required
async def check_kode(kode: str = Query(""), exclude_id: str = Query("")):
    available = barang_service.check_kode(kode, exclude_id)
    return {
        "available": available,
        "message": "Kode tersedia." if available else "Kode sudah digunakan.",
    }


@barang_bp.get("/low-stock")
@role_required("admin")
async def low_stock(limit: int = Query(20)):
    return {"data": barang_service.list_low_stock(limit=limit)}


@barang_bp.get("/lookup")
@api_login_required
async def lookup(keyword: str = Query(""), limit: int = Query(50)):
    return {"data": barang_service.list_lookup(keyword=keyword, limit=limit)}


@barang_bp.get("/import-template")
@role_required("admin")
async def import_template():
    """Generate and download XLSX template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import Barang"
    headers = ["kode_barang", "nama_barang", "nama_kategori", "satuan", "stok_awal", "stok_minimum", "lokasi_barang", "deskripsi_barang"]
    ws.append(headers)
    ws.append(["BRG-001", "Contoh Barang 1", "Elektronik", "pcs", 10, 2, "Gudang A", "Deskripsi contoh"])
    ws.append(["BRG-002", "Contoh Barang 2", "ATK", "pcs", 50, 10, "Rak 1", "Deskripsi barang 2"])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        body=buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template-import-barang.xlsx"'},
    )


@barang_bp.post("/import")
@role_required("admin")
async def import_xlsx(file: UploadFile = File(...)):
    """Import data barang dari XLSX."""
    if not file or not file.filename:
        raise HTTPBadRequest("File tidak valid.")
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPBadRequest("File harus .xlsx atau .xls")
    try:
        content = await file.read()
        wb = load_workbook(_io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPBadRequest("File kosong.")
        headers = [str(h or "").strip().lower() for h in rows[0]]
        result = barang_service.import_xlsx(headers, rows[1:])
        return result
    except ValueError as exc:
        raise HTTPBadRequest(str(exc))
    except Exception as exc:
        raise HTTPBadRequest(f"Gagal import: {exc}")


@barang_bp.post("/upload-foto")
@role_required("admin")
async def upload_foto(file: UploadFile = File(...), kode: str = Query("")):
    """Upload foto barang ke Cloudinary, kembalikan info foto."""
    if not kode:
        raise HTTPBadRequest("Parameter kode wajib diisi.")
    if not file or not file.filename:
        raise HTTPBadRequest("File tidak valid.")
    try:
        info = cloudinary_service.upload_barang_photo(
            file.file, kode, filename=file.filename
        )
    except Exception as exc:
        raise HTTPBadRequest(f"Gagal upload foto: {exc}")
    return info


@barang_bp.post("/upload-foto-base64")
@role_required("admin")
async def upload_foto_base64(kode: str = Query("")):
    """Upload foto barang via base64 (JSON). Fallback ketika multipart gagal."""
    import base64 as _b64
    import json as _json
    import sys as _sys
    _raw_body = request.body
    payload = {}
    if _raw_body:
        try:
            payload = _json.loads(_raw_body)
        except Exception:
            payload = request.json or {}
    kode = (kode or "").strip() or (payload.get("kode") or "").strip()
    filename = (payload.get("filename") or "photo").strip()
    content_type = (payload.get("content_type") or "application/octet-stream").strip()
    data_b64 = (payload.get("data") or "").strip()
    if not kode:
        raise HTTPBadRequest("Parameter kode wajib diisi.")
    if not data_b64:
        raw_head = _raw_body[:200].decode("utf-8", errors="replace") if _raw_body else "(empty)"
        _dbg = {
            "payload_keys": list(payload.keys()),
            "has_data": "data" in payload,
            "data_type": type(payload.get("data")).__name__ if "data" in payload else "MISSING",
            "raw_len": len(_raw_body) if _raw_body else 0,
            "raw_head": raw_head,
        }
        print(f"[UPLOAD-B64 EMPTY BODY] {_json.dumps(_dbg, default=str)}", file=_sys.stderr, flush=True)
        raise HTTPBadRequest("Data foto kosong.")
    try:
        raw = _b64.b64decode(data_b64, validate=True)
    except Exception as exc:
        raise HTTPBadRequest(f"Data base64 tidak valid: {exc}")
    try:
        ext = "jpg"
        if filename and "." in filename:
            ext = filename.rsplit(".", 1)[-1].lower()
        info = cloudinary_service.upload_barang_photo(
            _io.BytesIO(raw), kode, filename=filename, ext=ext
        )
    except Exception as exc:
        raise HTTPBadRequest(f"Gagal upload foto: {exc}")
    return info


@barang_bp.post("/upload-foto-raw")
@role_required("admin")
async def upload_foto_raw(kode: str = Query("")):
    """Upload foto via raw binary body — paling sederhana, hindari JSON/FormData."""
    raw = request.body
    if not raw:
        raise HTTPBadRequest("Data foto kosong.")
    content_type = request.headers.get("Content-Type", "application/octet-stream")
    ext = "jpg"
    ct_lower = content_type.lower()
    if "png" in ct_lower:
        ext = "png"
    elif "webp" in ct_lower:
        ext = "webp"
    elif "gif" in ct_lower:
        ext = "gif"
    filename = f"photo.{ext}"
    try:
        info = cloudinary_service.upload_barang_photo(
            _io.BytesIO(raw), kode, filename=filename, ext=ext
        )
    except Exception as exc:
        raise HTTPBadRequest(f"Gagal upload foto: {exc}")
    return info


# ---- setelah literal-path routes, baru param route ----

@barang_bp.get("/<barang_id>")
@api_login_required
async def show(barang_id: str):
    doc = barang_service.get_barang(barang_id)
    if not doc:
        raise HTTPNotFound("Barang tidak ditemukan.")
    return doc


@barang_bp.post("")
@role_required("admin")
async def create(payload: dict = Body(...)):
    try:
        return barang_service.create_barang(payload)
    except ValueError as exc:
        raise HTTPBadRequest(str(exc))


@barang_bp.put("/<barang_id>")
@role_required("admin")
async def update(barang_id: str, payload: dict = Body(...)):
    try:
        result = barang_service.update_barang(barang_id, payload)
    except ValueError as exc:
        raise HTTPBadRequest(str(exc))
    if not result:
        raise HTTPNotFound("Barang tidak ditemukan.")
    return result


@barang_bp.delete("/<barang_id>")
@role_required("admin")
async def destroy(barang_id: str):
    try:
        deleted = barang_service.delete_barang(barang_id)
    except ValueError as exc:
        raise HTTPBadRequest(str(exc))
    if not deleted:
        raise HTTPNotFound("Barang tidak ditemukan.")
    return {"message": "Barang berhasil dihapus."}


@barang_bp.get("/<barang_id>/riwayat-stok")
@api_login_required
async def riwayat_stok(barang_id: str):
    items = barang_service.get_riwayat_stok(barang_id)
    return {"data": items}
